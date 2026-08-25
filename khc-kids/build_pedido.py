# -*- coding: utf-8 -*-
"""
Pedido Inicial · Llenar la tienda KHC Kids 3×6 m
Importación desde China · calidad media · análisis de tallas y capacidades
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PINK="F9A8C4"; PINK_D="E36A98"; PINK_LT="FFE3EF"
BLUE="A9D8EA"; BLUE_D="4E96BF"; BLUE_LT="DCEEFC"
YEL="FFE19A"; YEL_D="E0A93A"; YEL_LT="FFF1CF"
GREEN="C7E5B5"; GREEN_D="5C8A3C"; GREEN_LT="E3F3E0"
PEACH="FFD0B0"; PEACH_D="D87A4A"; PEACH_LT="FFE3D3"
LILAC="D9C4EE"; LILAC_D="9170C4"; LILAC_LT="EDE0EF"
CREAM="FFFBF6"; INK="5A4852"; SOFT="9A8A92"; LINE="F3E5EC"; WHITE="FFFFFF"

CUR='"$"#,##0.00'; KG='0.00" kg"'; NUM='#,##0'; PCT='0%'
thin=Side(style="thin",color=LINE); border=Border(left=thin,right=thin,top=thin,bottom=thin)
def F(bold=False,size=11,color=INK): return Font(name="Calibri",bold=bold,size=size,color=color)
def fill(c): return PatternFill("solid",fgColor=c)
center=Alignment(horizontal="center",vertical="center",wrap_text=True)
left=Alignment(horizontal="left",vertical="center",wrap_text=True)
right=Alignment(horizontal="right",vertical="center")
def m(x): return float(x)

# ================= DATOS DEL PEDIDO =================
# cat, producto, und, precio china (mediana calidad), peso unit kg, precio venta USD
ITEMS=[
 # ---- BEBÉS 0-24m ----
 ("Bebés","Bodies manga corta (surtidos)",60,1.50,0.055,5.00),
 ("Bebés","Mamelucos cortos frescos",25,3.50,0.100,9.00),
 ("Bebés","Ajuares recién nacido 4 pz",10,7.00,0.280,14.00),
 ("Bebés","Conjuntos body + short",15,4.00,0.140,9.50),
 ("Bebés","Camisones frescos",10,3.00,0.110,7.50),
 ("Bebés","Baberos pack 5",10,2.50,0.150,7.00),
 # ---- NIÑAS 2-10 ----
 ("Niñas","Vestidos casuales frescos",45,4.50,0.120,9.50),
 ("Niñas","Vestidos de fiesta",12,6.50,0.180,15.00),
 ("Niñas","Conjuntos top + short",25,5.00,0.190,12.00),
 ("Niñas","Faldas plisadas",20,3.00,0.100,7.00),
 ("Niñas","Blusitas sin manga",20,2.20,0.070,4.50),
 ("Niñas","Palazzos / pantalones frescos",15,3.50,0.160,9.00),
 ("Niñas","Shorts niña",13,2.20,0.080,4.50),
 # ---- NIÑOS 2-10 ----
 ("Niños","Franelas manga corta",45,2.50,0.090,4.50),
 ("Niños","Conjuntos franela + short",25,5.00,0.190,12.00),
 ("Niños","Pantalones frescos jogger",25,3.50,0.160,9.00),
 ("Niños","Bermudas mezclillo",30,2.80,0.140,8.00),
 ("Niños","Polos piqué",25,2.80,0.100,7.00),
 # ---- PIJAMAS & UNISEX ----
 ("Pijamas","Pijamas cortos 2 pz",35,4.00,0.150,9.50),
 ("Pijamas","Pijamas largos ligeros 2 pz",20,4.50,0.180,10.00),
 ("Pijamas","Camisones niña",10,3.50,0.130,10.00),
 ("Pijamas","Caliszones pack 4",10,3.00,0.140,6.00),
 # ---- TRAJES DE BAÑO ----
 ("Baño","Trajes de baño niña",15,3.50,0.080,10.00),
 ("Baño","Bermudas de baño niño",15,3.00,0.090,9.50),
 # ---- CALZADO ----
 ("Calzado","Sandalias frescas",25,4.50,0.180,12.00),
 ("Calzado","Chanclas / slippers",15,2.50,0.120,6.00),
 # ---- ACCESORIOS ----
 ("Accesorios","Calcetines pack 5 pares",25,2.50,0.090,6.50),
 ("Accesorios","Gorras anti-sol",30,1.80,0.060,5.50),
 ("Accesorios","Gafas de sol UV",20,1.20,0.040,5.00),
 ("Accesorios","Mochilas infantiles",15,4.50,0.200,11.00),
 ("Accesorios","Sets coleteras / accesorios pelo",15,1.50,0.050,4.50),
]

tot_cost=sum(i[2]*i[3] for i in ITEMS)
tot_units=sum(i[2] for i in ITEMS)
tot_weight=sum(i[2]*i[4] for i in ITEMS)
revenue=sum(i[2]*i[5] for i in ITEMS)

# curvas de talla
CURVE_BEBE=[("0-3M",.10),("3-6M",.20),("6-12M",.30),("12-18M",.25),("18-24M",.15)]
CURVE_KIDS=[("2-3A",.18),("3-4A",.22),("4-5A",.22),("5-6A",.16),("6-7A",.12),("8-9A",.06),("10A",.04)]
CURVE_SHOE=[("18",.05),("20",.08),("22",.15),("24",.18),("26",.18),("28",.15),("30",.13),("32",.08)]

bebe_units=sum(i[2] for i in ITEMS if i[0]=="Bebés")
kids_units=sum(i[2] for i in ITEMS if i[0] in ("Niñas","Niños"))
shoe_units=sum(i[2] for i in ITEMS if i[0]=="Calzado")

# envío y costos
PACK=1.06  # 6% de empaques/cedulos
bill_weight=tot_weight*PACK
RATE_SEA=4.0; RATE_AIR=12.0
ship_sea=bill_weight*RATE_SEA; ship_air=bill_weight*RATE_AIR
SAMPLES=60.0; AGENT=150.0; CONTINGENCY=tot_cost*0.05
extra=SAMPLES+AGENT+CONTINGENCY
landed_sea=tot_cost+ship_sea+extra
landed_air=tot_cost+ship_air+extra
profit_sea=revenue-landed_sea
profit_air=revenue-landed_air

# versión ajustada 70%
sc=.70
adj_cost=tot_cost*sc; adj_weight=bill_weight*sc
adj_sea=adj_cost+adj_weight*RATE_SEA+(SAMPLES+AGENT+adj_cost*0.05)
adj_units=tot_units*sc; adj_rev=revenue*sc

wb=openpyxl.Workbook()

# ============ HOJA: PEDIDO ============
ws=wb.active; ws.title="Pedido"; ws.sheet_view.showGridLines=False
ws["A1"]="KHC KIDS · PEDIDO INICIAL PARA LLENAR LA TIENDA (3×6 m)"
ws["A1"].font=F(bold=True,size=18,color=PINK_D); ws.merge_cells("A1:H1")
ws["A2"]="Importación China · calidad media (algodón 140–180 g/m²) · precios mayoristas USD"
ws["A2"].font=F(color=SOFT); ws.merge_cells("A2:H2")
heads=["#","Categoría","Producto","Unidades","Precio China","Peso unit.","Peso total","Costo línea"]
for c,h in enumerate(heads,start=1):
    cell=ws.cell(4,c,h); cell.font=F(bold=True,color=WHITE); cell.fill=fill(PINK_D); cell.alignment=center; cell.border=border
r=5; cat_colors={"Bebés":PEACH_LT,"Niñas":PINK_LT,"Niños":BLUE_LT,"Pijamas":LILAC_LT,"Baño":BLUE_LT,"Calzado":PINK_LT,"Accesorios":GREEN_LT}
last_cat=None; subtotal_rows=[]
for n,it in enumerate(ITEMS,start=1):
    cat,name,und,cost,w,sell=it
    if last_cat and cat!=last_cat:
        sub=[i for i in ITEMS if i[0]==last_cat]
        ws.cell(r,3,f"Subtotal {last_cat}").font=F(bold=True,color=INK)
        ws.cell(r,4,sum(i[2] for i in sub)).number_format=NUM
        ws.cell(r,7,sum(i[2]*i[4] for i in sub)).number_format=KG
        ws.cell(r,8,sum(i[2]*i[3] for i in sub)).number_format=CUR
        for c in range(1,9): ws.cell(r,c).fill=fill(WHITE); ws.cell(r,c).border=border; ws.cell(r,c).font=F(bold=True,size=10)
        subtotal_rows.append(r); r+=1
    bg=cat_colors.get(cat,WHITE)
    vals=[n,cat,name,und,cost,w,und*w,und*cost]
    for c,v in enumerate(vals,start=1):
        cell=ws.cell(r,c,v); cell.border=border; cell.fill=fill(bg); cell.font=F()
        if c==5: cell.number_format=CUR
        if c==4: cell.number_format=NUM; cell.alignment=center
        if c in (6,7): cell.number_format='0.000" kg"' if c==6 else KG; cell.alignment=center
        if c==8: cell.number_format=CUR
        if c==1: cell.alignment=center
        if c==2: cell.font=F(bold=True,size=9); cell.alignment=center
        if c==3: cell.alignment=left
    last_cat=cat; r+=1
# subtotal último
sub=[i for i in ITEMS if i[0]==last_cat]
ws.cell(r,3,f"Subtotal {last_cat}").font=F(bold=True)
ws.cell(r,4,sum(i[2] for i in sub)).number_format=NUM
ws.cell(r,7,sum(i[2]*i[4] for i in sub)).number_format=KG
ws.cell(r,8,sum(i[2]*i[3] for i in sub)).number_format=CUR
for c in range(1,9): ws.cell(r,c).border=border; ws.cell(r,c).font=F(bold=True,size=10)
r+=1
# TOTAL
for c in range(1,9):
    cell=ws.cell(r,c); cell.fill=fill(PINK_D); cell.border=border; cell.font=F(bold=True,color=WHITE)
ws.cell(r,3,"TOTAL PEDIDO").font=F(bold=True,color=WHITE,size=12)
ws.cell(r,4,tot_units).font=F(bold=True,color=WHITE,size=12); ws.cell(r,4).number_format=NUM
ws.cell(r,7,round(tot_weight,2)).number_format=KG
ws.cell(r,8,tot_cost).number_format=CUR; ws.cell(r,8).font=F(bold=True,color=WHITE,size=12)
widths=[5,13,34,10,12,11,11,12]
for i,wv in enumerate(widths,start=1): ws.column_dimensions[get_column_letter(i)].width=wv
ws.freeze_panes="A5"

# ============ HOJA: TALLAS ============
wt=wb.create_sheet("Tallas"); wt.sheet_view.showGridLines=False
wt["A1"]="ANÁLISIS DE TALLAS · CURVAS DE SURTIDO"
wt["A1"].font=F(bold=True,size=18,color=PINK_D); wt.merge_cells("A1:F1")
wt["A2"]="Cómo repartir las unidades entre tallas según la demanda real de cada edad"
wt["A2"].font=F(color=SOFT); wt.merge_cells("A2:F2")

def curve_table(row,title,curve,units,accent,bglt,nota):
    wt.cell(row,1,title).font=F(bold=True,size=12,color=accent)
    row+=1
    heads=["Talla","% del lote","Unidades a pedir"]
    for c,h in enumerate(heads,start=1):
        cell=wt.cell(row,c,h); cell.font=F(bold=True,color=WHITE); cell.fill=fill(accent); cell.alignment=center; cell.border=border
    row+=1
    for t,p in curve:
        wt.cell(row,1,t).border=border; wt.cell(row,1).alignment=center; wt.cell(row,1).fill=fill(bglt)
        c2=wt.cell(row,2,p); c2.number_format=PCT; c2.border=border; c2.alignment=center; c2.fill=fill(bglt)
        c3=wt.cell(row,3,round(units*p)); c3.number_format=NUM; c3.border=border; c3.alignment=center; c3.fill=fill(bglt)
        row+=1
    wt.cell(row,1,"Total").font=F(bold=True)
    c3=wt.cell(row,3,units); c3.number_format=NUM; c3.font=F(bold=True); c3.alignment=center
    for c in (1,2,3): wt.cell(row,c).border=border
    row+=2
    wt.cell(row,1,nota).font=Font(name="Calibri",size=9,color=SOFT,italic=True); wt.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6)
    return row+2

row=4
row=curve_table(row,"👶 BEBÉS · 0–24 meses ({} unidades: bodies, mamelucos, conjuntos…)".format(bebe_units),CURVE_BEBE,bebe_units,PEACH_D,PEACH_LT,
  "La talla 6-12M es la más vendida: los bebés crecen rápido y los papás compran más en esta etapa. La 0-3M se vende menos (regalos de nacimiento).")
row=curve_table(row,"🧒 NIÑOS/AS · 2–10 años ({} unidades de prendas de niña y niño)".format(kids_units),CURVE_KIDS,kids_units,PINK_D,PINK_LT,
  "El grueso de la venta está en 3-5 años (talla de regalo y crecimiento). Las tallas 8-10 se piden pocas: menos público y los pre-adolescentes ya eligen tiendas de adulto.")
row=curve_table(row,"👟 CALZADO · tallas 18–32 ({} pares)".format(shoe_units),CURVE_SHOE,shoe_units,BLUE_D,BLUE_LT,
  "En calzado infantil el pico está en 22–28 (apenas los peques caminan y crecen cada 6 meses). Pide poco en 18-20 (bebitos no caminan) y 30-32.")

# mix del pedido
wt.cell(row,1,"MIX DEL PEDIDO POR CATEGORÍA").font=F(bold=True,size=13,color=PINK_D); row+=1
for c,h in enumerate(["Categoría","Unidades","% del pedido","Costo (USD)"],start=1):
    cell=wt.cell(row,c,h); cell.font=F(bold=True,color=WHITE); cell.fill=fill(PINK_D); cell.alignment=center; cell.border=border
row+=1
for cat in ["Bebés","Niñas","Niños","Pijamas","Baño","Calzado","Accesorios"]:
    sub=[i for i in ITEMS if i[0]==cat]
    u=sum(i[2] for i in sub); cst=sum(i[2]*i[3] for i in sub)
    vals=[cat,u,u/tot_units,cst]
    for c,v in enumerate(vals,start=1):
        cell=wt.cell(row,c,v); cell.border=border; cell.fill=fill(cat_colors.get(cat,WHITE))
        if c==2: cell.number_format=NUM; cell.alignment=center
        if c==3: cell.number_format=PCT; cell.alignment=center
        if c==4: cell.number_format=CUR
    row+=1
for c,v in enumerate(["TOTAL",tot_units,1.0,tot_cost],start=1):
    cell=wt.cell(row,c,v); cell.font=F(bold=True,color=WHITE); cell.fill=fill(PINK_D); cell.border=border
    if c==2: cell.number_format=NUM; cell.alignment=center
    if c==3: cell.number_format=PCT; cell.alignment=center
    if c==4: cell.number_format=CUR
row+=2
wt.cell(row,1,"NOTAS DE COMPRA EN CHINA (MOQ y surtidos)").font=F(bold=True,size=12,color=GREEN_D); row+=1
for note in [
 "• Muchos proveedores (Alibaba/1688) venden por 'lotes surtidos': pide que respeten TU curva de tallas o agrupa varios modelos en el mismo lote.",
 "• MOQ típico: 20–50 und por modelo/color. Con 31 modelos y ~22 und promedio es viable; negocia 'mixed sizes' en un mismo MOQ.",
 "• Pide MUESTRAS antes ($50–60): verifica gramaje del algodón (140–180 g/m²), costuras y que no destiña.",
 "• Calidad media = algodón 100% o mezcla 65/35, elástico en cinturas, etiquetas lavables. Evita telas muy finas que se deforman.",
 "• Pide etiquetas neutras y añade tu propia etiqueta KHC Kids (complemento de marca).",
 "• Escaparate: reserva 2 outfits completos de lo mejor del lote para los maniquíes.",
]:
    wt.cell(row,1,note).font=F(size=10); wt.merge_cells(start_row=row,start_column=1,end_row=row,end_column=6); row+=1
for col,wv in zip("ABCDEF",[28,12,16,14,12,12]): wt.column_dimensions[col].width=wv

# ============ HOJA: COSTOS ============
wc=wb.create_sheet("Costos"); wc.sheet_view.showGridLines=False
wc["A1"]="COSTOS TOTALES DEL PEDIDO (LANDED COST)"
wc["A1"].font=F(bold=True,size=18,color=PINK_D); wc.merge_cells("A1:D1")
wc["A2"]="Importación China → Venezuela · compara marítimo vs aéreo"
wc["A2"].font=F(color=SOFT); wc.merge_cells("A2:D2")
def kv(row,label,value,fmt=CUR,bold=False,color=INK,bg=None,note=""):
    a=wc.cell(row,1,label); a.font=F(bold=bold,color=color); a.alignment=left
    b=wc.cell(row,2,value); b.font=F(bold=bold,color=color); b.number_format=fmt; b.alignment=right
    if bg: a.fill=fill(bg); b.fill=fill(bg)
    if note: c=wc.cell(row,3,note); c.font=F(size=9,color=SOFT)
    return row+1
r=4
r=kv(r,"Productos (31 modelos)",tot_cost,note=f"{tot_units} unidades · calidad media")
r=kv(r,"Peso neto",tot_weight,KG)
r=kv(r,"Peso facturable (con empaques +6%)",bill_weight,KG,bold=True)
r+=1
wc.cell(r,1,"OPCIÓN A · MARÍTIMO (recomendada)").font=F(bold=True,size=12,color=GREEN_D); r+=1
r=kv(r,"Flete marítimo ($4/kg × 60–70 días)",ship_sea)
r=kv(r,"Muestras previas",SAMPLES,note="verificar calidad antes del lote")
r=kv(r,"Gestión/agente de compra",AGENT,note="sourcing, consolidación, documentos")
r=kv(r,"Imprevistos (5%)",CONTINGENCY)
r=kv(r,"LANDED COST MARÍTIMO",landed_sea,bold=True,color=GREEN_D,bg=GREEN_LT)
r=kv(r,"Costo por prenda",landed_sea/tot_units,bold=True,note=f"entre {tot_units} und")
r+=1
wc.cell(r,1,"OPCIÓN B · AÉREO (solo si tienes prisa)").font=F(bold=True,size=12,color=YEL_D); r+=1
r=kv(r,"Flete aéreo ($12/kg × 7–12 días)",ship_air)
r=kv(r,"Extras (muestras + agente + 5%)",extra)
r=kv(r,"LANDED COST AÉREO",landed_air,bold=True,color=YEL_D,bg=YEL_LT)
r=kv(r,"Costo por prenda",landed_air/tot_units,bold=True)
r+=2
wc.cell(r,1,"VERSIÓN AJUSTADA (si el capital es justo)").font=F(bold=True,size=12,color=PINK_D); r+=1
r=kv(r,"70% del pedido (productos)",adj_cost,note=f"~{adj_units:.0f} unidades")
r=kv(r,"Landed marítimo (70%)",adj_sea,bold=True,bg=PINK_LT)
r=kv(r,"Ingreso potencial (70%)",adj_rev)
wc.column_dimensions["A"].width=36; wc.column_dimensions["B"].width=16; wc.column_dimensions["C"].width=42

# ============ HOJA: VENTA ============
wv=wb.create_sheet("Venta"); wv.sheet_view.showGridLines=False
wv["A1"]="VENTA Y GANANCIA POR PRODUCTO"
wv["A1"].font=F(bold=True,size=18,color=PINK_D); wv.merge_cells("A1:G1")
wv["A2"]="Precios de venta de tu tienda KHC Kids (consistente con el catálogo) · ganancia con landed marítimo"
wv["A2"].font=F(color=SOFT); wv.merge_cells("A2:G2")
heads=["Producto","Und.","Costo China","Peso total","Costo landed líneaa","Ingreso","Ganancia"]
# corregir encabezado
heads=["Producto","Und.","Costo línea","Ingreso","Costo landed (marít.)","Ganancia","Margen"]
for c,h in enumerate(heads,start=1):
    cell=wv.cell(4,c,h); cell.font=F(bold=True,color=WHITE); cell.fill=fill(PINK_D); cell.alignment=center; cell.border=border
sea_per_kg=(ship_sea+extra)/bill_weight
r=5
for it in ITEMS:
    cat,name,und,cost,w,sell=it
    line_cost=und*cost; line_w=und*w
    landed_line=line_cost+line_w*PACK*sea_per_kg
    income=und*sell; gan=income-landed_line
    vals=[name,und,line_cost,income,landed_line,gan,(gan/income if income else 0)]
    for c,v in enumerate(vals,start=1):
        cell=wv.cell(r,c,v); cell.border=border; cell.fill=fill(cat_colors.get(cat,WHITE)); cell.font=F(size=10)
        if c==2: cell.number_format=NUM; cell.alignment=center
        if c in (3,4,5,6): cell.number_format=CUR
        if c==7: cell.number_format=PCT; cell.alignment=center
        if c==1: cell.alignment=left
    r+=1
for c,v in enumerate(["TOTAL",tot_units,tot_cost,revenue,landed_sea,profit_sea,(profit_sea/revenue)],start=1):
    cell=wv.cell(r,c,v); cell.font=F(bold=True,color=WHITE); cell.fill=fill(GREEN_D); cell.border=border
    if c==2: cell.number_format=NUM; cell.alignment=center
    if c in (3,4,5,6): cell.number_format=CUR
    if c==7: cell.number_format=PCT; cell.alignment=center
widths=[32,8,11,11,15,11,9]
for i,wv2 in enumerate(widths,start=1): wv.column_dimensions[get_column_letter(i)].width=wv2
wv.freeze_panes="A5"

# ============ HOJA: RESUMEN (primera) ============
wr=wb.create_sheet("Resumen",0); wr.sheet_view.showGridLines=False
wr["B2"]="KHC KIDS · PEDIDO INICIAL DE APERTURA"
wr["B2"].font=F(bold=True,size=22,color=PINK_D)
wr["B3"]="Llenar la tienda física 3×6 m · Importación China calidad media · USD"
wr["B3"].font=F(size=12,color=SOFT)
def card(r,c,title,value,sub,color,bg):
    wr.cell(r,c,title).font=F(bold=True,size=10,color=SOFT)
    vc=wr.cell(r+1,c,value); vc.font=F(bold=True,size=20,color=color)
    wr.cell(r+2,c,sub).font=F(size=9,color=SOFT)
    for rr in range(r,r+3):
        for cc in (c,c+1):
            wr.cell(rr,cc).fill=fill(bg); wr.cell(rr,cc).border=Border(left=Side(style="medium",color=bg),right=Side(style="medium",color=bg))
card(5,2,"UNIDADES DEL PEDIDO",tot_units,"31 modelos · 7 categorías",PINK_D,PINK_LT)
card(5,4,"PRODUCTOS (FOB CHINA)",tot_cost,f"promedio ${tot_cost/tot_units:.2f}/prenda",BLUE_D,BLUE_LT)
card(5,6,"INVERSIÓN TOTAL (MARÍTIMO)",landed_sea,"productos + flete + gestión",GREEN_D,GREEN_LT)
card(5,8,"GANANCIA SI VENDES TODO",profit_sea,f"{profit_sea/landed_sea*100:.0f}% ROI · {profit_sea/revenue*100:.0f}% margen",GREEN_D,GREEN_LT)

wr["B10"]="FICHA RÁPIDA DEL PEDIDO"; wr["B10"].font=F(bold=True,size=13,color=PINK_D)
rows=[
 ("Modelos (referencias)",31),("Unidades totales",tot_units),
 ("Peso facturable",f"{bill_weight:.1f} kg"),("Costo FOB China",f"${tot_cost:,.2f}"),
 ("Costo promedio por prenda",f"${tot_cost/tot_units:.2f}"),
 ("Landed marítimo (total)",f"${landed_sea:,.2f}"),("Landed marítimo por prenda",f"${landed_sea/tot_units:.2f}"),
 ("Ingreso si vendes todo",f"${revenue:,.2f}"),("Ganancia (marítimo)",f"${profit_sea:,.2f}"),
 ("ROI marítimo",f"{profit_sea/landed_sea*100:.0f}%"),("Ganancia (aéreo)",f"${profit_air:,.2f}"),
]
r=11
for k,v in rows:
    wr.cell(r,2,k).font=F(bold=(r%2==0))
    c=wr.cell(r,3,v); c.font=F(bold=True,color=(GREEN_D if "Ganancia" in k or "ROI" in k else INK)); c.alignment=right
    bg=WHITE if r%2 else CREAM
    for cc in (2,3): wr.cell(r,cc).fill=fill(bg)
    r+=1
r+=1
wr.cell(r,2,"CAPACIDAD DE LA TIENDA vs PEDIDO").font=F(bold=True,size=13,color=PINK_D); r+=1
for c,h in enumerate(["Zona (del plano)","Capacidad aprox.","Pedido","Cobertura"],start=2):
    cell=wr.cell(r,c,h); cell.font=F(bold=True,color=WHITE); cell.fill=fill(PINK_D); cell.alignment=center; cell.border=border
r+=1
cap=[("Niñas (barra doble + repisas)",180,150),("Niños (barra doble + repisas)",180,150),
     ("Bebés (cuna + estantería baja)",120,130),("Pijamas & unisex (barra + cubos)",115,75),
     ("Mesas centrales (ofertas/combos)",120,120),("Repisas fondo (accesorios)",70,65),
     ("Calzado (cubos bajo mesas)",50,40),("Escaparate (maniquíes)",4,4)]
for z,capv,ordv in cap:
    vals=[z,capv,ordv,(ordv/capv if capv else 0)]
    for c,v in enumerate(vals,start=2):
        cell=wr.cell(r,c,v); cell.border=border
        if c in (3,4): cell.number_format=NUM; cell.alignment=center
        if c==5: cell.number_format=PCT; cell.alignment=center
    r+=1
r+=1
tips=[
 f"1) Inversión total de apertura: ${landed_sea:,.0f} de mercancía + ~$2.750 de acondicionamiento ≈ ${landed_sea+2750:,.0f}.",
 "2) Pide por curva de tallas (hoja Tallas): la 6-12M y la 3-5A son las más vendidas; 8-10 años solo el 10%.",
 "3) El pedido deja ~20% de las prendas como stock de rotación bajo el mostrador: la tienda SIEMPRE se ve llena.",
 "4) Flete marítimo ($4/kg) casi triplica tu ganancia frente al aéreo ($12/kg). Solo usa aéreo para reponer 'best sellers'.",
 "5) Pide muestras antes de todo el lote ($50–60) y verifica algodón 140–180 g/m² y costuras.",
 "6) Escaparate: reserva los 2 outfits más llamativos para los maniquíes y cámbialos cada 2 semanas.",
]
wr.cell(r,2,"CONSEJOS CLAVE").font=F(bold=True,size=13,color=PINK_D); r+=1
for t in tips:
    wr.cell(r,2,t).font=F(size=10); wr.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9); r+=1
wr.column_dimensions["A"].width=2
for col in "BCDEFGHI": wr.column_dimensions[col].width=17
wr.column_dimensions["B"].width=34

# ============ HOJA: MARCA PROPIA ============
wm=wb.create_sheet("Marca"); wm.sheet_view.showGridLines=False
wm["A1"]="MARCA PROPIA (KHC KIDS) · ¿CUÁNTO CUESTA?"
wm["A1"].font=F(bold=True,size=18,color=PINK_D); wm.merge_cells("A1:E1")
wm["A2"]="Precios reales de proveedores chinos (Made-in-China / Alibaba, 2025-2026) · MOQ 500-1.000 etiquetas"
wm["A2"].font=F(color=SOFT); wm.merge_cells("A2:E2")

# escenarios
LBL_COST=0.04; LBL_QTY=1000
SEW=0.05
TAG=0.05; TAG_QTY=700
BAG=0.06; BAG_QTY=700
escA=LBL_COST*LBL_QTY+SEW*tot_units
escB=escA+TAG*TAG_QTY+BAG*BAG_QTY
rev_up=revenue*1.10
landed_B=landed_sea+escB
gan_B=rev_up-landed_B

r=4
wm.cell(r,1,"LOS 3 CAMINOS PARA VENDER CON TU MARCA").font=F(bold=True,size=13,color=PINK_D); r+=1
for c,h in enumerate(["Camino","En qué consiste","Coste extra aprox.","¿Para ti?"],start=1):
    cell=wm.cell(r,c,h); cell.font=F(bold=True,color=WHITE); cell.fill=fill(PINK_D); cell.alignment=center; cell.border=border
r+=1
caminos=[
 ("A · Etiqueta tejida cosida","Compras el stock sin marca (o retiran la original) y cosen TU etiqueta 'KHC Kids' en el cuello",
  f"${escA:,.0f} (${escA/tot_units:.2f}/prenda)","✅ Lo esencial · buenísimo"),
 ("B · Marca completa","Etiqueta tejida + hang tag colgante con tu logo + bolsa poly personalizada",
  f"${escB:,.0f} (${escB/tot_units:.2f}/prenda)","⭐ RECOMENDADO"),
 ("C · Fabricación OEM total","Fábrica la prenda desde cero con tu marca, etiquetas y empaques",
  "+20-35% por prenda y MOQ 50-300 por modelo","⛔ No para el 1er pedido"),
]
for c1,c2,c3,c4 in caminos:
    for c,v in enumerate([c1,c2,c3,c4],start=1):
        cell=wm.cell(r,c,v); cell.border=border; cell.alignment=left if c==2 else center; cell.font=F(size=10,bold=(c==1))
    r+=1
r+=1
wm.cell(r,1,"DESGLOSE CON PRECIOS REALES (Escenario B · recomendado)").font=F(bold=True,size=13,color=PINK_D); r+=1
for c,h in enumerate(["Concepto","Cantidad","Precio und.","Total"],start=1):
    cell=wm.cell(r,c,h); cell.font=F(bold=True,color=WHITE); cell.fill=fill(BLUE_D); cell.alignment=center; cell.border=border
r+=1
detalle=[
 ("Etiquetas tejidas 'KHC Kids' (diseño tuyo, damasco)",LBL_QTY,LBL_COST),
 ("Cosido/retiro de etiqueta original en fábrica",tot_units,SEW),
 ("Hang tags colgantes con logo + cordel",TAG_QTY,TAG),
 ("Bolsas poly personalizadas (empaque por prenda)",BAG_QTY,BAG),
]
for n,q,p in detalle:
    vals=[n,q,p,q*p]
    for c,v in enumerate(vals,start=1):
        cell=wm.cell(r,c,v); cell.border=border
        if c==2: cell.number_format=NUM; cell.alignment=center
        if c==3: cell.number_format='"$"#,##0.000'; cell.alignment=center
        if c==4: cell.number_format=CUR
        if c==1: cell.alignment=left; cell.font=F(size=10)
    r+=1
for c,v in enumerate(["TOTAL EXTRA","","",escB],start=1):
    cell=wm.cell(r,c,v); cell.font=F(bold=True,color=WHITE); cell.fill=fill(GREEN_D); cell.border=border
    if c==4: cell.number_format=CUR
r+=2
wm.cell(r,1,"IMPACTO EN TU PEDIDO DE {} PRENDAS".format(tot_units)).font=F(bold=True,size=13,color=PINK_D); r+=1
for c,h in enumerate(["Concepto","Sin marca","Con marca (B)","Diferencia"],start=1):
    cell=wm.cell(r,c,h); cell.font=F(bold=True,color=WHITE); cell.fill=fill(PINK_D); cell.alignment=center; cell.border=border
r+=1
impact=[
 ("Costo productos (FOB)",tot_cost,tot_cost+escB),
 ("Inversión total (landed marítimo)",landed_sea,landed_B),
 ("Costo promedio por prenda",tot_cost/tot_units,(tot_cost+escB)/tot_units),
 ("Ingreso si vendes todo (precios actuales)",revenue,revenue),
 ("Ingreso si subes precios +10% (con marca)",revenue,rev_up),
 ("GANANCIA final",profit_sea,gan_B),
]
for k,v1,v2 in impact:
    vals=[k,v1,v2,v2-v1]
    for c,v in enumerate(vals,start=1):
        cell=wm.cell(r,c,v); cell.border=border
        if c>1: cell.number_format=CUR
        if c==1: cell.alignment=left; cell.font=F(size=10)
        if c==4: cell.font=F(size=10,bold=True,color=(GREEN_D if v2>v1 else SOFT))
        if k.startswith("GANANCIA"): cell.font=F(bold=True,color=WHITE if c>0 else INK); cell.fill=fill(GREEN_LT)
    r+=1
r+=1
for tip in [
 "1) MOQ de etiquetas es 500-1.000: pide 1.000 y te sobran ~320 para reponer en el 2do pedido.",
 "2) Muchos proveedores de ropa cosen la etiqueta GRATIS si compras el lote: negégialo ('free re-labeling service').",
 "3) Con etiqueta propia el cliente no puede googlear 'precio Shein': deja de comparar y acepta tu precio.",
 "4) Ropa con marca se percibe 10-20% más valiosa: +10% en tus precios ya paga el branding 2-3 veces.",
 "5) Pide MUESTRA de la etiqueta tejida antes de producirlas (verifica colores del logo y suavidad).",
 "6) El diseño de tu etiqueta: fondo blanco/rosa, 'KHC KIDS' + 'Hecho con amor' + composición y talla.",
 "7) Tiempo extra: +3 a 7 días de producción por el cosido de etiquetas.",
]:
    wm.cell(r,1,tip).font=F(size=10); wm.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5); r+=1
for col,wv in zip("ABCDE",[42,14,14,14,14]): wm.column_dimensions[col].width=wv

wb.save("/home/user/khc-kids/Pedido-Inicial-Tienda-KHC-Kids.xlsx")
print("✅ Excel generado")
print(f"Modelos: 31 | Unidades: {tot_units} | Peso facturable: {bill_weight:.1f} kg")
print(f"FOB China: ${tot_cost:,.2f} (promedio ${tot_cost/tot_units:.2f}/und)")
print(f"Landed MARÍTIMO: ${landed_sea:,.2f} → ganancia ${profit_sea:,.2f} ({profit_sea/landed_sea*100:.0f}% ROI)")
print(f"Landed AÉREO:   ${landed_air:,.2f} → ganancia ${profit_air:,.2f} ({profit_air/landed_air*100:.0f}% ROI)")
print(f"Ingreso total: ${revenue:,.2f}")
print(f"Mix: Bebés {bebe_units} | Niñas/Niños {kids_units} | Calzado {shoe_units}")
