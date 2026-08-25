# -*- coding: utf-8 -*-
"""
Plan de Negocio KHC Kids · Comprar en Shein y revender en Venezuela
Genera un libro de Excel exacto y editable.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- PALETA pastel ----------
PINK   = "F8AFC4"; PINK_D = "E36A98"; PINK_LT = "FFE3EF"
BLUE   = "A9D8EA"; BLUE_LT = "DCEEFC"
YELLOW = "FFE19A"; YEL_LT = "FFF1CF"
GREEN  = "C7E5B5"; GRN_LT = "E3F3E0"
LILAC  = "D9C4EE"; LIL_LT = "EDE0EF"
CREAM  = "FFFBF6"; INK = "5A4852"; SOFT = "9A8A92"; LINE = "F3E5EC"
WHITE  = "FFFFFF"

CUR = '"$"#,##0.00'
KG  = '0.00" kg"'
NUM = '#,##0'
PCT = '0.0%'

thin = Side(style="thin", color=LINE)
med  = Side(style="medium", color="D8BFCB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def F(bold=False, size=11, color=INK, name="Calibri"):
    return Font(name=name, bold=bold, size=size, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left", vertical="center", wrap_text=True)
right  = Alignment(horizontal="right", vertical="center")

# ---------- DATOS ----------
# cat, prenda, edad, precio_shein(lote), qty(lotes), unid/lote, peso_lote(kg), unidades_a_vender, precio_venta(USD)
items = [
 # ---- BEBÉS 0-24m (todo fresco, sin pies ni abrigo) ----
 ("Bebés","Pack 5 Bodies manga corta algodón","0-24m",9.50,10,5,0.28,50,5.00),
 ("Bebés","Mameluco / pelele CORTO fresco (pack 2)","0-24m",8.00,8,2,0.22,16,9.00),
 ("Bebés","Conjunto fresco body + short bebé","6-24m",6.50,10,1,0.16,10,9.50),
 ("Bebés","Vestido fresco de verano bebé niña","6-24m",4.50,16,1,0.11,16,8.50),
 ("Bebés","Camisón / enterizo fresco de algodón","0-24m",4.99,8,1,0.12,8,7.50),
 ("Bebés","Gorra anti-sol bebé","0-24m",3.50,8,1,0.06,8,5.50),
 # ---- NIÑA fresca 2-10a ----
 ("Niña","Vestido fresco niña casual","2-10a",5.99,16,1,0.13,16,9.50),
 ("Niña","Conjunto top + short fresco niña","2-7a",7.99,10,1,0.20,10,12.00),
 ("Niña","Blusitas / sin manga niña (pack 3)","2-10a",8.99,6,3,0.22,18,4.50),
 ("Niña","Vestido fiesta fresco niña","3-10a",8.99,6,1,0.18,6,15.00),
 ("Niña","Leggings delgados niña (pack 3)","2-10a",8.50,6,3,0.28,18,4.50),
 ("Niña","Pantalón largo fresco lino/algodón niña","2-10a",5.50,10,1,0.18,10,9.00),
 ("Niña","Traje de baño / bikini niña","2-10a",5.99,10,1,0.08,10,10.00),
 ("Niña","Short / pantaloncito niña (pack 3)","2-7a",7.99,6,3,0.24,18,4.50),
 # ---- NIÑO fresco 2-10a ----
 ("Niño","Franelas manga corta niño (pack 3)","2-10a",9.99,6,3,0.28,18,4.50),
 ("Niño","Conjunto franela + short fresco niño","2-7a",7.99,10,1,0.20,10,12.00),
 ("Niño","Pantalón largo fresco jogger ligero niño","2-10a",5.50,10,1,0.18,10,9.00),
 ("Niño","Short deportivo niño (pack 3)","2-10a",8.99,6,3,0.26,18,4.50),
 ("Niño","Traje de baño / bermuda baño niño","2-10a",5.50,10,1,0.09,10,9.50),
 # ---- PIJAMAS FRESCOS ----
 ("Pijama fresco","Pijama corto camiseta+short 2pz","2-10a",5.99,12,1,0.16,12,9.50),
 ("Pijama fresco","Pijama pantalón largo ligero 2pz","2-10a",6.50,10,1,0.19,10,10.00),
 # ---- ACCESORIOS & CALZADO FRESCO ----
 ("Accesorios","Gorra / sombrero anti-sol (pack 2)","Todas",4.50,10,2,0.08,20,5.50),
 ("Accesorios","Sandalias frescas bebé / niño","Todas",6.99,8,1,0.18,8,12.00),
 ("Accesorios","Calcetines finos (pack 5 pares)","Todas",4.99,8,1,0.08,8,6.50),
]

# totales
tot_cost = sum(i[3]*i[4] for i in items)
tot_units = sum(i[7] for i in items)
tot_weight = sum(i[4]*i[6] for i in items)
tot_revenue = sum(i[7]*i[8] for i in items)

RATE_AIR = 12.0     # USD por kg aéreo (Zoom/Liberty/Alas Latinas ~12 USD/kg)
RATE_SEA = 4.0      # USD por kg marítimo (Tramvene/Liberty ~3.5-5 USD/kg)
EXTRA = 80.0        # DUA + seguro parcial + empaques/bolsas
ship_air = round(tot_weight*RATE_AIR,2)
ship_sea = round(tot_weight*RATE_SEA,2)
landed_air = round(tot_cost+ship_air+EXTRA,2)
landed_sea = round(tot_cost+ship_sea+EXTRA,2)
profit_air = round(tot_revenue-landed_air,2)
profit_sea = round(tot_revenue-landed_sea,2)
unit_air = round(landed_air/tot_units,2)
unit_sea = round(landed_sea/tot_units,2)

wb = openpyxl.Workbook()

# =====================================================================
# HOJA 1 — LISTA DE COMPRA
# =====================================================================
ws = wb.active
ws.title = "Compra"
ws.sheet_view.showGridLines = False
ws["A1"] = "KHC KIDS · Lista de Compra en SHEIN"
ws["A1"].font = F(bold=True, size=18, color=PINK_D)
ws.merge_cells("A1:K1")
ws["A2"] = "Selección 100% TROPICAL · prendas frescas, sin abrigo · Presupuesto máx $1.500 (solo productos) · USD"
ws["A2"].font = F(color=SOFT); ws.merge_cells("A2:K2")

headers = ["#","Categoría","Prenda","Edad","Precio Shein\n(USD/lote)","Cant.\nlotes","Unid.\n/lote","Unidades\ntotales","Peso lote\n(kg)","Peso total\n(kg)","Costo línea\n(USD)"]
hrow = 4
for c,h in enumerate(headers, start=1):
    cell = ws.cell(hrow, c, h)
    cell.font = F(bold=True, color=WHITE)
    cell.fill = fill(PINK_D)
    cell.alignment = center
    cell.border = border
ws.row_dimensions[hrow].height = 34

cat_colors = {"Bebés":"FFE3D3","Niña":"FFE3EF","Niño":"DCEEFC","Pijama fresco":"FFF1CF","Accesorios":"E3F3E0"}
r = hrow+1
for n,it in enumerate(items, start=1):
    cat,name,age,price,qty,per,w,usell,psell = it
    units = qty*per
    weight = qty*w
    cost = price*qty
    row = [n,cat,name,age,price,qty,per,units,w,weight,cost]
    bg = cat_colors.get(cat, WHITE)
    for c,val in enumerate(row, start=1):
        cell = ws.cell(r,c,val)
        cell.border = border
        cell.fill = fill(bg)
        cell.font = F()
        if c in (5,11): cell.number_format = CUR
        if c in (6,7,8): cell.number_format = NUM; cell.alignment = center
        if c in (9,10): cell.number_format = KG; cell.alignment = center
        if c==1: cell.alignment = center
        if c==2: cell.font = F(bold=True, size=9); cell.alignment = center
        if c==3: cell.alignment = left
        if c==4: cell.alignment = center; cell.font = F(size=9, color=SOFT)
    r += 1

# total row
tr = r
ws.cell(tr,1,"").fill = fill(PINK_D)
for c in range(1,12):
    cell = ws.cell(tr,c); cell.fill = fill(PINK_D); cell.border = border
ws.cell(tr,3,"TOTAL").font = F(bold=True, color=WHITE, size=12); ws.cell(tr,3).alignment = right
ws.cell(tr,8,tot_units).font = F(bold=True, color=WHITE); ws.cell(tr,8).number_format=NUM; ws.cell(tr,8).alignment=center
ws.cell(tr,10,tot_weight).font = F(bold=True, color=WHITE); ws.cell(tr,10).number_format=KG; ws.cell(tr,10).alignment=center
ws.cell(tr,11,tot_cost).font = F(bold=True, color=WHITE, size=12); ws.cell(tr,11).number_format=CUR

# presupuesto check
ws.cell(tr+2,3,"Presupuesto usado:").font = F(bold=True)
ws.cell(tr+2,5,tot_cost).number_format=CUR; ws.cell(tr+2,5).font=F(bold=True, color=PINK_D)
ws.cell(tr+3,3,"Presupuesto máximo:").font = F(bold=True)
ws.cell(tr+3,5,1500).number_format=CUR
ws.cell(tr+4,3,"Disponible aún:").font = F(bold=True)
ws.cell(tr+4,5,1500-tot_cost).number_format=CUR; ws.cell(tr+4,5).font=F(bold=True, color="3C9D5A")

widths = [5,12,32,9,12,7,7,10,10,11,12]
for i,wv in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = wv
ws.freeze_panes = "A5"

# =====================================================================
# HOJA 2 — ENVÍO Y COSTOS
# =====================================================================
we = wb.create_sheet("Envio")
we.sheet_view.showGridLines = False
we["A1"] = "Costos de Envío y Landed Cost (a Venezuela)"
we["A1"].font = F(bold=True, size=18, color=PINK_D); we.merge_cells("A1:D1")
we["A2"] = "Comparamos las 2 opciones reales de envío · Todos los valores en USD"
we["A2"].font = F(color=SOFT); we.merge_cells("A2:D2")

def kv(row, label, value, fmt=CUR, bold=False, color=INK, bg=None, note=""):
    a = we.cell(row,1,label); a.font = F(bold=bold, color=color); a.alignment=left
    b = we.cell(row,2,value); b.font = F(bold=bold, color=color); b.number_format=fmt; b.alignment=right
    if bg:
        a.fill=fill(bg); b.fill=fill(bg)
    if note:
        c = we.cell(row,3,note); c.font = F(size=9, color=SOFT); c.alignment=left
    return b

we.cell(4,1,"DATOS DEL PEDIDO").font = F(bold=True, color=PINK_D, size=12)
kv(5,"Total en productos (Shein)",tot_cost, note="= presupuesto usado")
kv(6,"Peso total estimado",tot_weight,KG, note="ropa + accesorios empacados")
kv(7,"Unidades totales",tot_units,NUM)

we.cell(9,1,"OPCIÓN A · ENVÍO AÉREO (rápido)").font = F(bold=True, color="2E7BA6", size=12)
kv(10,"Tarifa aérea (USD/kg)",RATE_AIR,note="Zoom / Liberty / Alas Latinas ≈ $12/kg")
kv(11,"Costo envío aéreo",ship_air, note=f"{tot_weight} kg × $12 (se factura kilo completo)")
we.cell(12,1,"OPCIÓN B · ENVÍO MARÍTIMO (barato)").font = F(bold=True, color="5C8A3C", size=12)
kv(13,"Tarifa marítimo (USD/kg)",RATE_SEA,note="Tramvene / Liberty / cajas ≈ $3,5–5/kg")
kv(14,"Costo envío marítimo",ship_sea, note="más lento: 35–60 días")

we.cell(16,1,"GASTOS ADICIONALES").font = F(bold=True, color=PINK_D, size=12)
kv(17,"DUA + seguro + empaques/bolsas",EXTRA, note="aduanas, cinta, bolsas, etiquetas")

we.cell(19,1,"LANDED COST (costo total de importar)").font = F(bold=True, color=PINK_D, size=12)
kv(20,"Landed cost AÉREO",landed_air,bold=True,bg=BLUE_LT)
kv(21,"Landed cost MARÍTIMO",landed_sea,bold=True,bg=GRN_LT)
kv(22,"Costo por prenda AÉREO",unit_air,note=f"{tot_units} prendas")
kv(23,"Costo por prenda MARÍTIMO",unit_sea,note=f"{tot_units} prendas")

we.column_dimensions["A"].width = 34
we.column_dimensions["B"].width = 16
we.column_dimensions["C"].width = 46

# =====================================================================
# HOJA 3 — PRECIOS DE VENTA Y GANANCIA
# =====================================================================
wv = wb.create_sheet("Venta")
wv.sheet_view.showGridLines = False
wv["A1"] = "Precios de Venta en Venezuela y Ganancia"
wv["A1"].font = F(bold=True, size=18, color=PINK_D); wv.merge_cells("A1:K1")
wv["A2"] = "Selección tropical · Precios de venta (USD) según MercadoLibre VE · Ganancia comparada por tipo de envío"
wv["A2"].font = F(color=SOFT); wv.merge_cells("A2:K2")

heads = ["#","Prenda","Und.\nvender","Precio\nventa","Ingreso\ntotal","Costo\ncompra","Gan.\nAÉREO","Gan.\nMARÍTIMO","Margen\nAÉREO","Margen\nMARÍTIMO","⭐"]
hr = 4
for c,h in enumerate(heads, start=1):
    cell = wv.cell(hr,c,h); cell.font=F(bold=True,color=WHITE); cell.fill=fill(PINK_D); cell.alignment=center; cell.border=border
wv.row_dimensions[hr].height = 34

# reparto de envío por peso
share_air_per_kg = (ship_air+EXTRA)/tot_weight
share_sea_per_kg = (ship_sea+EXTRA)/tot_weight

r = hr+1
for n,it in enumerate(items, start=1):
    cat,name,age,price,qty,per,w,usell,psell = it
    cost = price*qty
    weight = qty*w
    revenue = usell*psell
    gan_air = revenue - cost - weight*share_air_per_kg
    gan_sea = revenue - cost - weight*share_sea_per_kg
    mar_air = gan_air/revenue if revenue else 0
    mar_sea = gan_sea/revenue if revenue else 0
    star = "★★★" if mar_sea>=0.45 else ("★★" if mar_sea>=0.30 else "★")
    vals = [n,name,usell,psell,revenue,cost,gan_air,gan_sea,mar_air,mar_sea,star]
    bg = cat_colors.get(cat, WHITE)
    for c,val in enumerate(vals, start=1):
        cell = wv.cell(r,c,val); cell.border=border; cell.fill=fill(bg); cell.font=F()
        if c in (4,5,6,7,8): cell.number_format=CUR
        if c==3: cell.number_format=NUM; cell.alignment=center
        if c in (9,10): cell.number_format=PCT; cell.alignment=center
        if c==1: cell.alignment=center
        if c==2: cell.alignment=left
        if c==11: cell.alignment=center; cell.font=F(color="E0A93A")
    r += 1

tr=r
for c in range(1,12):
    cell=wv.cell(tr,c); cell.fill=fill(PINK_D); cell.border=border
wv.cell(tr,2,"TOTAL").font=F(bold=True,color=WHITE,size=12); wv.cell(tr,2).alignment=right
wv.cell(tr,5,tot_revenue).font=F(bold=True,color=WHITE); wv.cell(tr,5).number_format=CUR
wv.cell(tr,6,tot_cost).font=F(bold=True,color=WHITE); wv.cell(tr,6).number_format=CUR
wv.cell(tr,7,profit_air).font=F(bold=True,color=WHITE); wv.cell(tr,7).number_format=CUR
wv.cell(tr,8,profit_sea).font=F(bold=True,color=WHITE); wv.cell(tr,8).number_format=CUR
wv.cell(tr,9,profit_air/tot_revenue).font=F(bold=True,color=WHITE); wv.cell(tr,9).number_format=PCT; wv.cell(tr,9).alignment=center
wv.cell(tr,10,profit_sea/tot_revenue).font=F(bold=True,color=WHITE); wv.cell(tr,10).number_format=PCT; wv.cell(tr,10).alignment=center

widths=[5,30,8,9,11,11,11,12,9,10,6]
for i,wv2 in enumerate(widths, start=1):
    wv.column_dimensions[get_column_letter(i)].width=wv2
wv.freeze_panes="A5"

# =====================================================================
# HOJA 4 — RESUMEN
# =====================================================================
wr = wb.create_sheet("Resumen", 0)  # primera
wr.sheet_view.showGridLines=False
wr["B2"]="KHC KIDS · Resumen del Plan de Negocio"
wr["B2"].font=F(bold=True,size=22,color=PINK_D)
wr["B3"]="Comprar en Shein y revender en Venezuela · Selección 100% TROPICAL · Niños 0–10 años"
wr["B3"].font=F(size=12,color=SOFT)

def card(r, c, title, value, fmt, color, bg, sub=""):
    wr.cell(r,c,title).font=F(bold=True,color=SOFT,size=10)
    wr.cell(r,c).fill=fill(bg)
    wr.cell(r,c).alignment=Alignment(horizontal="left",vertical="center")
    vc = wr.cell(r+1,c,value); vc.font=F(bold=True,size=20,color=color); vc.number_format=fmt
    vc.fill=fill(bg); vc.alignment=Alignment(horizontal="left",vertical="center")
    if sub:
        wr.cell(r+2,c,sub).font=F(size=9,color=SOFT); wr.cell(r+2,c).fill=fill(bg)
    for rr in range(r,r+3):
        wr.cell(rr,c).fill=fill(bg); wr.cell(rr,c).border=Border(left=med,top=med,bottom=med)
        wr.cell(rr,c+1).border=Border(right=med,top=med,bottom=med); wr.cell(rr,c+1).fill=fill(bg)
    wr.merge_cells(start_row=r,start_column=c,end_row=r+2,end_column=c+1)

# tarjetas superiores
card(5,2,"INVERSIÓN EN PRODUCTOS",tot_cost,CUR,PINK_D,PINK_LT,f"{tot_units} prendas · {tot_weight:.0f} kg")
card(5,4,"INGRESO SI VENDES TODO",tot_revenue,CUR,"2E7BA6",BLUE_LT,"precios mercado VE")
card(5,6,"GANANCIA · MARÍTIMO",profit_sea,CUR,"5C8A3C",GRN_LT,f"{profit_sea/tot_revenue*100:.0f}% margen · {profit_sea/landed_sea*100:.0f}% ROI")
card(5,8,"GANANCIA · AÉREO",profit_air,CUR,"C77E20",YEL_LT,f"{profit_air/tot_revenue*100:.0f}% margen · {profit_air/landed_air*100:.0f}% ROI")

# tabla comparativa
wr["B10"]="COMPARATIVA DE LAS DOS OPCIONES DE ENVÍO"
wr["B10"].font=F(bold=True,color=PINK_D,size=13)
comp = [
 ("Concepto","AÉREO","MARÍTIMO"),
 ("Costo productos (Shein)",tot_cost,tot_cost),
 ("Peso del envío",f"{tot_weight:.2f} kg",f"{tot_weight:.2f} kg"),
 ("Tarifa por kg","$12,00/kg","$4,00/kg"),
 ("Costo de envío",ship_air,ship_sea),
 ("Gastos extra (DUA+seguro+empaques)",EXTRA,EXTRA),
 ("LANDED COST (total importado)",landed_air,landed_sea),
 ("Ingreso si vendes todo",tot_revenue,tot_revenue),
 ("GANANCIA TOTAL",profit_air,profit_sea),
 ("Margen sobre venta",f"{profit_air/tot_revenue*100:.1f}%",f"{profit_sea/tot_revenue*100:.1f}%"),
 ("ROI (sobre lo invertido)",f"{profit_air/landed_air*100:.0f}%",f"{profit_sea/landed_sea*100:.0f}%"),
 ("Costo por prenda importada",unit_air,unit_sea),
 ("Tiempo de entrega","7 – 12 días","35 – 60 días"),
 ("Recomendado para","lotes pequeños / pruebas","VOLUMEN y ganancia"),
]
start=11
for i,(a,b,c) in enumerate(comp):
    rr=start+i
    bg = WHITE if i%2 else CREAM
    if i==0: bg=PINK_D
    if a=="GANANCIA TOTAL": bg=GREEN
    ca=wr.cell(rr,2,a); cb=wr.cell(rr,3,b); cc=wr.cell(rr,4,c)
    for cell in (ca,cb,cc):
        cell.border=border; cell.fill=fill(bg)
        cell.font=F(bold=(i==0 or a in("GANANCIA TOTAL","LANDED COST (total importado)","ROI (sobre lo invertido)")),
                    color=WHITE if i==0 else (INK))
        cell.alignment=left if cell is ca else right
    # formatos numéricos
    if isinstance(b,(int,float)) and i not in (0,2,3,12,13):
        cb.number_format=CUR; cc.number_format=CUR

wr.column_dimensions["A"].width=2
wr.column_dimensions["B"].width=38
for col in ("C","D"): wr.column_dimensions[col].width=22
for col in ("E","F","G","H","I"): wr.column_dimensions[col].width=16

# conclusión
cr=start+len(comp)+2
wr.cell(cr,2,"💡 CONCLUSIÓN Y CONSEJOS").font=F(bold=True,color=PINK_D,size=13)
tips = [
 f"1) Usa ENVÍO MARÍTIMO para el volumen: casi DUPLICA tu ganancia ({profit_sea/landed_sea*100:.0f}% ROI vs {profit_air/landed_air*100:.0f}% aéreo). Reserva el aéreo solo para pruebas o los más vendidos.",
 "2) Como TODO es tropical y ligero, el envío pesa menos (≈36 kg) y el margen por prenda SUBE respecto a un mix con abrigos/zapatos pesados.",
 "3) Vende las prendas baratas SUELTAS (bodies $5, leggings y shorts $4,50, blusitas $4,50) y los multipacks como paquete: más unidades, precio accesible.",
 f"4) Con $1.500 traes {tot_units} prendas frescas. Si vendes todo: +${profit_sea:,.0f} de ganancia (marítimo). Realista: descuenta 15–20% que no se vende al inicio.",
 "5) Plan SIN abrigos ni sudaderas: pantalones largos FINOS (lino/algodón/jogger ligero) y trajes de baño, ideales para clima tropical. 90% de las prendas bajo $13.",
 "6) Empaca en bolsas con tu marca KHC Kids + etiqueta de talla: suma valor y permite cobrar un poco más.",
 "7) Aduana VE: por casillero (Zoom/Tealca) suelen no cobrar arancel en envíos personales pequeños; declara valor moderado y reparte en 1–2 guías.",
]
for i,t in enumerate(tips):
    cell=wr.cell(cr+1+i,2,t); cell.font=F(); cell.alignment=left
    wr.merge_cells(start_row=cr+1+i,start_column=2,end_row=cr+1+i,end_column=8)

wb.save("/home/user/khc-kids/Plan-de-Negocio-KHC-Kids.xlsx")
print("✅ Excel generado")
print(f"   Productos: ${tot_cost:.2f}  |  Prendas: {tot_units}  |  Peso: {tot_weight:.2f} kg")
print(f"   Ingreso: ${tot_revenue:.2f}")
print(f"   Landed AÉREO: ${landed_air:.2f} → Ganancia ${profit_air:.2f} ({profit_air/tot_revenue*100:.1f}% margen, {profit_air/landed_air*100:.0f}% ROI)")
print(f"   Landed MARÍTIMO: ${landed_sea:.2f} → Ganancia ${profit_sea:.2f} ({profit_sea/tot_revenue*100:.1f}% margen, {profit_sea/landed_sea*100:.0f}% ROI)")
